from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active

# switch values
altitude = "altitude"
humidity = "Humidity"
CO2 = "CO2: "
CO = "The concentration of CO is "
NO2 = "NO2"
CH4 = "The concentration of CH4 is "

sheet["A1"] = "Altitude(meters)"
sheet["F1"] = "Humidity %"
sheet["B1"] = "CO2 (ppm)"
sheet["C1"] = "CO (ppm)"
sheet["D1"] = "NO2 (ppm)"
sheet["E1"] = "CH4 (ppm)"

a_row = 2
col = 1
a_cell = ""
DataFile = open("DATAFILE.TXT")
for line in DataFile:
    final_data = ""
    if altitude in line:
        for n in line:
            if n.isdigit():
                final_data += n
            if n == ".":
                final_data += n
        col = 1
        sheet.cell(row=a_row, column=col, value=None).value = final_data
    if "Humidity" in line:
        for n in line:
            if n.isdigit():
                final_data += n
            if n == ".":
                final_data += n
        col = 6
        sheet.cell(row=a_row, column=col, value=None).value = final_data
    if CO2 in line:
        for n in line:
            if n.isdigit():
                final_data += n
            if n == ".":
                final_data += n
        col = 2
        sheet.cell(row=a_row, column=col, value=None).value = final_data
    if CO in line:
        for n in line:
            if n.isdigit():
                final_data += n
            if n == ".":
                final_data += n
        col = 3
        sheet.cell(row=a_row, column=col, value=None).value = final_data
    if NO2 in line:
        for n in line:
            if n.isdigit():
                final_data += n
            if n == ".":
                final_data += n
        col = 4
        sheet.cell(row=a_row, column=col, value=None).value = final_data
    if CH4 in line:
        for n in line:
            if n.isdigit():
                final_data += n
            if n == ".":
                final_data += n
        col = 5
        sheet.cell(row=a_row, column=col, value=None).value = final_data
        a_row += 1

workbook.save("Analyze_Data.xlsx")
DataFile.close()
